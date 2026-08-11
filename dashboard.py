from __future__ import annotations

import json
import os
import socket
import unicodedata
import base64
from copy import deepcopy
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.error import URLError
from urllib.request import Request, urlopen
from urllib.parse import urlparse

import openpyxl

import sensum_sql


BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = BASE_DIR / "COMPROMISSO MES.xlsx"
PYTHON_EXCEL_PATH = BASE_DIR / "BASE PYTHON.xlsx"
LEGACY_EXCEL_PATH = BASE_DIR / "base_vendas.xlsx"
DATA_PATH = BASE_DIR / "data.json"
GERAL_DATA_PATH = BASE_DIR / "geral.json"
POSITIVACAO_MILHO_DATA_PATH = BASE_DIR / "positivacao_milho.json"
KEYS_DATA_PATH = BASE_DIR / "keys.json"
WEEKLY_GOALS_PATH = BASE_DIR / "weekly_goals.json"
LOGO_PATH = BASE_DIR / "iatagam_site_logo_cor.png"
HOST = "0.0.0.0"
PORT = int(os.environ.get("PORT", "8000"))
REMOTE_RAW_BASE = os.environ.get(
    "DASHBOARD_RAW_BASE",
    "https://raw.githubusercontent.com/JoaoCanhadas/acompanhamento-diario/main",
).rstrip("/")
REMOTE_API_BASE = os.environ.get(
    "DASHBOARD_API_BASE",
    "https://api.github.com/repos/JoaoCanhadas/acompanhamento-diario/contents",
).rstrip("/")
REMOTE_CACHE_TTL_SECONDS = int(os.environ.get("DASHBOARD_REMOTE_CACHE_SECONDS", "55"))
REMOTE_CACHE = {}
DEFAULT_WEEKLY_GOAL = 700000.0
POSITIVACAO_MILHO_WEEKLY_GOAL = 200.0
DEPLOY_VERSION = "2026-06-25 08:39"
WEEKDAY_LABELS = {
    0: "SEG",
    1: "TER",
    2: "QUA",
    3: "QUI",
    4: "SEX",
}


INDEX_HTML = r"""<!doctype html>
<html lang="pt-BR">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Acompanhamento Diario</title>
  <style>
    :root {
      color-scheme: dark;
      --bg: #05070d;
      --panel: #0c111d;
      --panel-2: #111827;
      --line: #223047;
      --ink: #f4f7fb;
      --soft: #a8b3c7;
      --muted: #79869d;
      --blue: #38bdf8;
      --blue-2: #2563eb;
      --purple: #8b5cf6;
      --good: #22c55e;
      --bad: #fb7185;
      --warn: #fbbf24;
      --shadow: 0 22px 60px rgba(0,0,0,.38);
    }

    * { box-sizing: border-box; }
    body {
      margin: 0;
      min-height: 100vh;
      font-family: "Segoe UI", Arial, sans-serif;
      color: var(--ink);
      background:
        radial-gradient(circle at 20% -10%, rgba(56,189,248,.18), transparent 30%),
        radial-gradient(circle at 90% 0%, rgba(139,92,246,.18), transparent 34%),
        var(--bg);
      overflow-x: hidden;
    }

    .topbar {
      border-bottom: 1px solid var(--line);
      background: rgba(5, 7, 13, .9);
      backdrop-filter: blur(18px);
    }

    .topbar-inner {
      max-width: 1520px;
      margin: 0 auto;
      padding: 20px 28px;
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 18px;
      flex-wrap: wrap;
    }

    .brand-logo {
      width: clamp(320px, 25vw, 440px);
      min-width: 320px;
      height: 76px;
      display: block;
    }

    .brand-logo img {
      display: block;
      width: 100%;
      height: 100%;
      object-fit: contain;
      object-position: left center;
    }

    .meta {
      display: flex;
      gap: 10px;
      flex-wrap: wrap;
      justify-content: flex-end;
      color: var(--soft);
      font-size: 14px;
      font-weight: 700;
    }

    .pill {
      padding: 9px 12px;
      border: 1px solid var(--line);
      border-radius: 999px;
      background: rgba(17, 24, 39, .8);
    }

    .view-tabs {
      display: flex;
      gap: 8px;
      padding: 4px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: rgba(8, 13, 22, .9);
    }

    .view-tab {
      height: 36px;
      padding: 0 14px;
      border: 0;
      border-radius: 6px;
      color: var(--soft);
      background: transparent;
      font-size: 14px;
      font-weight: 900;
      cursor: pointer;
    }

    .view-tab:not(.active):hover {
      background: rgba(255,255,255,.07);
      color: var(--ink);
    }

    .view-tab.active {
      color: var(--ink);
      background: linear-gradient(135deg, var(--blue-2), var(--purple));
    }

    .rotation-bar {
      height: 3px;
      background: rgba(34,48,71,.6);
      overflow: hidden;
    }

    .rotation-bar-fill {
      height: 100%;
      width: 0%;
      background: linear-gradient(90deg, var(--good), #84cc16);
    }

    .pause-btn {
      height: 36px;
      padding: 0 14px;
      border-radius: 6px;
      font-size: 13px;
      font-weight: 900;
      cursor: pointer;
      background: rgba(17,24,39,.8);
      border: 1px solid var(--line);
      color: var(--soft);
    }

    .pause-btn:hover {
      color: var(--ink);
      background: rgba(34,48,71,.9);
    }

    .pause-btn.paused {
      color: var(--warn);
      border-color: rgba(251,191,36,.4);
    }

    main {
      max-width: 1520px;
      margin: 0 auto;
      padding: 24px 28px 34px;
    }

    .cards {
  display: grid;
  grid-template-columns: repeat(4, minmax(0, 1fr));
  gap: 16px;
}

    .card, .panel {
      border: 1px solid var(--line);
      border-radius: 8px;
      background: linear-gradient(180deg, rgba(17,24,39,.96), rgba(12,17,29,.96));
      box-shadow: var(--shadow);
    }

    .card {
      min-height: 142px;
      padding: 20px;
      position: relative;
      overflow: hidden;
    }

    .card::after {
      content: "";
      position: absolute;
      inset: auto 0 0;
      height: 3px;
      background: linear-gradient(90deg, var(--blue), var(--purple));
    }

    .label {
      color: var(--soft);
      font-size: 15px;
      font-weight: 800;
      text-transform: uppercase;
    }

    .value {
  margin-top: 10px;
  font-size: 34px;
  line-height: 1.08;
  font-weight: 900;
  word-break: break-word;
}

/* VALOR DA PROJEÇÃO */
.projection-value {
  color: #FFC857;
}

/* LINHA DIVISÓRIA DO CARD PROJEÇÃO */
.projection-divider {
  height: 1px;
  margin: 6px 0 6px;
  background: rgba(148, 163, 184, 0.28);
}

/* DIFERENÇA PARA A META */
.projection-difference {
  margin-top: 4px;
  font-weight: 900;
}

.hint {
  margin-top: 10px;
  color: var(--muted);
  font-size: 14px;
  font-weight: 700;
}

    .good { color: var(--good); }
    .bad { color: var(--bad); }
    .accent { color: var(--blue); }

    .progress {
      width: 100%;
      height: 12px;
      margin-top: 16px;
      overflow: hidden;
      border-radius: 999px;
      background: #1b2638;
    }

    .progress span {
      display: block;
      height: 100%;
      width: 0;
      max-width: 100%;
      border-radius: inherit;
      background: linear-gradient(90deg, var(--good), #16a34a);
    }

    .layout {
      display: grid;
      grid-template-columns: minmax(0, 1.7fr) minmax(340px, .78fr);
      gap: 18px;
      align-items: start;
    }

    .panel-header {
      padding: 18px 20px;
      border-bottom: 1px solid var(--line);
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 12px;
      flex-wrap: wrap;
    }

    h2 {
      margin: 0;
      font-size: 20px;
      line-height: 1.2;
      font-weight: 900;
    }

    .controls {
      display: flex;
      gap: 10px;
      flex-wrap: wrap;
    }

    input, select, button {
      height: 42px;
      border: 1px solid var(--line);
      border-radius: 6px;
      background: #080d16;
      color: var(--ink);
      font: inherit;
      font-size: 15px;
      font-weight: 700;
      padding: 0 12px;
    }

    button {
      cursor: pointer;
      border-color: transparent;
      background: linear-gradient(135deg, var(--blue-2), var(--purple));
    }

    .table-wrap {
  height: 560px;
  max-height: 560px;
  overflow: hidden;
  position: relative;
}

    table {
      width: 100%;
      border-collapse: collapse;
      min-width: 860px;
      table-layout: fixed;
    }

    th, td {
      padding: 14px 16px;
      border-bottom: 1px solid rgba(34,48,71,.8);
      text-align: right;
      white-space: nowrap;
      font-size: 16px;
    }

    th:first-child, td:first-child { text-align: left; }
    th {
      color: var(--soft);
      background: rgba(8,13,22,.7);
      font-size: 13px;
      text-transform: uppercase;
      cursor: pointer;
      user-select: none;
      position: sticky;
      top: 0;
      z-index: 2;
    }

    thead {
  display: table;
  width: 100%;
  table-layout: fixed;
}

   thead th {
  position: relative;
  z-index: 20;
  background: #0c111d;
}

    tbody {
      display: block;
      will-change: transform;
    }

    tbody tr {
      display: table;
      width: 100%;
      table-layout: fixed;
    }

    .loop-copy {
      opacity: .96;
    }

    tbody tr:hover { background: rgba(56,189,248,.06); }
    .seller {
      color: #ffffff;
      font-weight: 900;
    }

    .money { font-weight: 900; }

    .tag {
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-width: 92px;
      height: 30px;
      border-radius: 999px;
      font-size: 13px;
      font-weight: 900;
      border: 1px solid;
    }

    .tag.ok {
      color: #8ef8b7;
      background: rgba(34,197,94,.12);
      border-color: rgba(34,197,94,.38);
    }

    .tag.pending {
      color: #ff9aad;
      background: rgba(251,113,133,.12);
      border-color: rgba(251,113,133,.38);
    }

    .weekly-wrap {
      height: 560px;
      max-height: 560px;
      overflow: hidden;
      position: relative;
    }

    .weekly {
      display: grid;
      gap: 14px;
      padding: 18px;
      will-change: transform;
    }

    .week-card {
      border: 1px solid var(--line);
      border-radius: 8px;
      background: rgba(8,13,22,.76);
      padding: 16px;
    }

    .week-top {
      display: flex;
      align-items: baseline;
      justify-content: space-between;
      gap: 12px;
      margin-bottom: 12px;
    }

    .week-name {
      font-size: 17px;
      font-weight: 900;
    }

    .week-percent {
      color: var(--blue);
      font-size: 22px;
      font-weight: 900;
    }

    .week-values {
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 10px;
      color: var(--soft);
      font-size: 14px;
      font-weight: 800;
    }

    .week-values strong {
      display: block;
      margin-top: 3px;
      color: var(--ink);
      font-size: 18px;
    }

    .error {
      margin-bottom: 18px;
      padding: 16px 18px;
      border-radius: 8px;
      border: 1px solid rgba(251,113,133,.5);
      color: #fecdd3;
      background: rgba(127,29,29,.4);
      font-weight: 800;
    }

    @media (max-width: 1180px) {
      .cards { grid-template-columns: repeat(2, minmax(0, 1fr)); }
      .layout { grid-template-columns: 1fr; }
    }

    @media (max-width: 640px) {
      body {
        background: var(--bg);
      }

      .topbar-inner, main {
        padding-left: 14px;
        padding-right: 14px;
      }

      .topbar-inner {
        align-items: flex-start;
        gap: 14px;
      }

      .brand-logo {
        width: min(300px, 100%);
        height: 66px;
        min-width: 0;
      }

      .meta {
        width: 100%;
        justify-content: flex-start;
      }

      .pill {
        max-width: 100%;
        overflow: hidden;
        text-overflow: ellipsis;
        white-space: nowrap;
      }

      main {
        padding-top: 16px;
        padding-bottom: 24px;
      }

      .cards {
        grid-template-columns: repeat(2, minmax(0, 1fr));
        gap: 10px;
      }

      .card {
        min-height: 118px;
        padding: 14px;
      }

      .label {
        font-size: 12px;
      }

      .value {
        font-size: 22px;
        line-height: 1.12;
      }

      .hint {
        font-size: 12px;
      }

      .progress {
        height: 9px;
        margin-top: 10px;
      }

      .layout {
        gap: 14px;
      }

      .panel-header {
        padding: 8px 12px;
        align-items: flex-start;
      }

      h2 {
        font-size: 16px;
      }

      .controls {
        width: 100%;
        display: grid;
        grid-template-columns: 1fr 1fr;
      }

      .controls input {
        grid-column: 1 / -1;
      }

      .controls input,
      .controls select,
      .controls button {
        width: 100%;
      }

      .table-wrap {
  height: 520px;
  max-height: 520px;
  overflow-y: auto;
  overflow-x: hidden;
}

      thead {
        display: none;
      }

      table,
      tbody,
      tr,
      td {
        display: block;
        width: 100%;
        min-width: 0;
      }

      tr.loop-copy {
        display: none;
      }

      tbody tr {
        margin: 12px;
        padding: 12px;
        border: 1px solid var(--line);
        border-radius: 8px;
        background: rgba(8,13,22,.76);
      }

      td {
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 12px;
        padding: 8px 0;
        border-bottom: 1px solid rgba(34,48,71,.5);
        text-align: right;
        white-space: normal;
        font-size: 14px;
      }

      td:last-child {
        border-bottom: 0;
      }

      td::before {
        content: attr(data-label);
        color: var(--soft);
        font-size: 12px;
        font-weight: 900;
        text-transform: uppercase;
        text-align: left;
      }

      td.seller {
        display: block;
        padding-top: 0;
        color: var(--ink);
        font-size: 16px;
        text-align: left;
      }

      td.seller::before {
        display: none;
      }

      .tag {
        min-width: 76px;
        height: 28px;
      }

      .weekly {
        padding: 14px;
      }

      .weekly-wrap {
        height: auto;
        max-height: none;
        overflow: visible;
      }

      .week-card {
        padding: 14px;
      }

      .week-values strong {
        font-size: 16px;
      }
    }

  </style>
</head>
<body>
  <header class="topbar">
    <div class="topbar-inner">
      <div class="brand-logo">
        <img src="/logo_iatagam_word.svg?v=3" alt="IATAGAM">
      </div>
      <div class="meta">
        <div class="view-tabs" aria-label="Telas do acompanhamento">
          <button class="view-tab active" data-view="sales" type="button">Varejo</button>
          <button class="view-tab" data-view="general" type="button">Geral</button>
          <button class="view-tab" data-view="keys" type="button">Key</button>
          <button class="view-tab" data-view="milho" type="button">Positivação</button>
        </div>
        <span class="pill" id="workbook">Carregando...</span>
        <span class="pill" id="updated"></span>
      </div>
    </div>
  </header>
  <main>
    <div id="error"></div>

    <section class="cards">
      <article class="card">
        <div class="label" id="commitmentLabel">Compromisso</div>
        <div class="value" id="commitment">R$ 0,00</div>
      </article>

      <article class="card">
        <div class="label" id="reachedLabel">Atingido</div>
        <div class="value good" id="reached">R$ 0,00</div>
        <div class="progress"><span id="totalProgress"></span></div>
      </article>

      <article class="card">
        <div class="label">Falta</div>
        <div class="value bad" id="missing">R$ 0,00</div>
        <div class="hint" id="pendingCount">0 vendedores pendentes</div>
      </article>

      <article class="card">
        <div class="label">Realizado</div>
        <div class="value accent" id="percent">0%</div>
        <div class="hint" id="positiveCount">0 acima da meta</div>
      </article>

      <article class="card projection-card" id="projectionCard">
        <div class="label">Projeção Mês</div>

        <div class="value projection-value" id="monthProjection">
          R$ 0,00
        </div>

        <div class="projection-divider"></div>

        <div class="hint" id="projectionDays">
          0 de 0 dias úteis
        </div>

        <div class="projection-difference" id="projectionDifference">
          R$ 0,00
        </div>
      </article>
    </section>

    <section class="layout">
      <article class="panel">
        <div class="panel-header">
          <h2 id="tableTitle">Performance Varejo</h2>
          <div class="controls">
            <input id="search" type="search" placeholder="Buscar vendedor">
            <select id="status">
              <option value="all">Todos</option>
              <option value="pending">Com falta</option>
              <option value="ok">Superou</option>
            </select>
            <button id="refresh" type="button">Atualizar</button>
          </div>
        </div>
        <div class="table-wrap">
          <table>
            <thead id="tableHead"></thead>
            <tbody id="tableBody"></tbody>
          </table>
        </div>
      </article>

      <aside class="panel">
        <div class="panel-header">
          <h2>Acompanhamento semanal</h2>
          <span class="pill" id="weekCount">4 semanas</span>
        </div>
        <div class="weekly-wrap">
          <div class="weekly" id="weekly"></div>
        </div>
      </aside>
    </section>
  </main>

  <script>
  const state = { rows: [], weeks: [], sortKey: "missing", sortDir: "desc" };

  let activeView = "sales";
  let loadRequestId = 0;
  let sellerScrollFrame = null;
  let weeklyScrollFrame = null;
  let sellerScrollOffset = 0;

  const brl = new Intl.NumberFormat("pt-BR", {
  style: "currency",
  currency: "BRL",
  minimumFractionDigits: 0,
  maximumFractionDigits: 0
});

const numberFormatter = new Intl.NumberFormat("pt-BR", {
  minimumFractionDigits: 0,
  maximumFractionDigits: 0
});

const byId = (id) => document.getElementById(id);

const formatMoney = (value) => {
  const number = Number(value || 0);
  return brl.format(Math.trunc(number));
};

const formatMoneyCompact = (value) => {
  const number = Math.abs(Number(value || 0));

  return `R$ ${Math.trunc(number).toLocaleString("pt-BR")}`;
};

const formatNumber = (value) => {
  const number = Number(value || 0);
  return numberFormatter.format(Math.trunc(number));
};

const formatBalance = (value, formatter) => {
  const amount = Number(value || 0);

  if (amount > 0) return formatter(-amount);
  if (amount < 0) return `+${formatter(Math.abs(amount))}`;

  return formatter(0);
};

  const moneyClass = (value) =>
    Number(value) <= 0 ? "good" : "bad";

  const setText = (id, value) => {
    byId(id).textContent = value;
  };

  const views = {
    sales: {
      endpoint: "/api/data",
      title: "Performance Varejo por vendedor",
      commitmentLabel: "Compromisso",
      pendingText: "vendedores pendentes",
      positiveText: "acima da meta",
      format: formatMoney,
      columns: [
        {
          key: "seller",
          label: "Vendedor",
          value: (row) => row.seller
        },
        {
          key: "commitment",
          label: "Compromisso",
          value: (row) => formatMoney(row.commitment)
        },
        {
          key: "reached",
          label: "Atingido",
          value: (row) => formatMoney(row.reached)
        },
        {
          key: "missing",
          label: "Falta",
          value: (row) => formatBalance(row.missing, formatMoney),
          className: (row) => `money ${moneyClass(row.missing)}`
        },
        {
          key: "percent",
          label: "%",
          value: (row) => `${row.percent}%`
        },
      ],
    },
      general: {
        endpoint: "/api/geral",
        title: "Performance Geral por vendedor",
        commitmentLabel: "Meta mes",
        pendingText: "vendedores pendentes",
        positiveText: "acima da meta",
        format: formatMoney,
        columns: [
  	{ key: "seller", label: "Vendedor", value: (row) => row.seller },
  	{ key: "commitment", label: "Meta", value: (row) => formatMoney(row.commitment) },
  	{ key: "reached", label: "Atingido", value: (row) => formatMoney(row.reached) },
  	{ key: "missing", label: "Falta", value: (row) => formatMoney(row.missing), className: (row) => `money ${moneyClass(row.missing)}` },
  	{ key: "percent", label: "%", value: (row) => `${Number(row.percent || 0).toFixed(1)}%` },
      ],
      },
      milho: {

  endpoint: "/api/positivacao-milho",

  title: "Performance Positivação por vendedor",

  commitmentLabel: "Meta mes",

  pendingText: "vendedores pendentes",

  positiveText: "bateram a meta",

  format: formatNumber,

  columns: [

    { key: "seller", label: "Vendedor", value: (row) => row.seller },

    { key: "commitment", label: "Meta", value: (row) => formatNumber(row.commitment) },

    { key: "reached", label: "Atingido", value: (row) => formatNumber(row.reached) },

    { key: "missing", label: "Falta", value: (row) => formatBalance(row.missing, formatNumber), className: (row) => `money ${moneyClass(row.missing)}` },

    { key: "percent", label: "%", value: (row) => `${Number(row.percent || 0).toFixed(1)}%` },

  ],

},
      keys: {
        endpoint: "/api/keys",
        title: "Performance Key por vendedor",
        commitmentLabel: "Meta semanal",
        pendingText: "vendedores pendentes",
        positiveText: "superaram a meta",
        format: formatMoney,
        columns: [
          { key: "seller", label: "Vendedor", value: (row) => row.seller },
          { key: "commitment", label: "Compromisso", value: (row) => formatMoney(row.commitment) },
          { key: "reached", label: "Atingido", value: (row) => formatMoney(row.reached) },
          { key: "missing", label: "Falta", value: (row) => formatBalance(row.missing, formatMoney), className: (row) => `money ${moneyClass(row.missing)}` },
          { key: "percent", label: "%", value: (row) => `${row.percent}%` },
        ],
      },
    };

    async function loadData() {
  const requestId = ++loadRequestId;
  const viewForRequest = activeView;
  byId("error").innerHTML = "";

  if (!state.rows.length) {
    renderLoading(viewForRequest);
  }

  try {
        const config = views[viewForRequest];
        const response = await fetch(config.endpoint + "?ts=" + Date.now());
        if (!response.ok) throw new Error(await response.text());
        const data = await response.json();
        if (requestId !== loadRequestId || viewForRequest !== activeView) return;
        state.rows = data.rows || [];
        state.weeks = normalizeWeeks(data);
        renderSummary(data, viewForRequest);
        renderTable(viewForRequest);
        renderWeeks(viewForRequest);
      } catch (error) {
        if (requestId !== loadRequestId || viewForRequest !== activeView) return;
        byId("error").innerHTML = `<div class="error">${error.message}</div>`;
      }
    }

    function renderLoading(viewName) {
  const config = views[viewName];

  const projectionCard = byId("projectionCard");
const cardsContainer = document.querySelector(".cards");

if (viewName === "general") {
  projectionCard.style.display = "";
  cardsContainer.style.gridTemplateColumns =
  "1.25fr 1.25fr 1.70fr 0.70fr 0.90fr";

  const resultado = calcularProjecaoMes(summary.reached);

  const diferenca =
    resultado.projecao - Number(summary.commitment || 0);

  setText(
    "monthProjection",
    formatMoneyCompact(resultado.projecao)
  );

  setText(
    "projectionDays",
    `${resultado.diasUteisDecorridos} de ${resultado.diasUteisMes} dias úteis`
  );

  if (diferenca >= 0) {
    setText(
      "projectionDifference",
      `+${formatMoney(diferenca)} (acima da meta)`
    );

    byId("projectionDifference").className =
      "projection-difference good";
  } else {
    setText(
      "projectionDifference",
      `-${formatMoney(Math.abs(diferenca))} (abaixo da meta)`
    );

    byId("projectionDifference").className =
      "projection-difference bad";
  }

} else {
  projectionCard.style.display = "none";
  cardsContainer.style.gridTemplateColumns =
    "repeat(4, minmax(0, 1fr))";
}

  setText(
    "workbook",
    `Carregando ${config.commitmentLabel.toLowerCase()}...`
  );

  setText("updated", "");
  setText("tableTitle", config.title);
  setText("commitmentLabel", config.commitmentLabel);
  setText("commitment", config.format(0));
  setText("reached", config.format(0));
  setText("missing", config.format(0));

  byId("missing").className = "value";

  setText("percent", "0%");
  setText("pendingCount", "");
  setText("positiveCount", "");

  byId("totalProgress").style.width = "0%";

  renderTableHead(config);

  byId("tableBody").innerHTML =
    `<tr><td data-label="Carregando" colspan="${config.columns.length + 1}">Carregando...</td></tr>`;

  byId("weekly").innerHTML =
    `<div class="week-card">Carregando...</div>`;
}

function normalizeWeeks(data) {
  if (Array.isArray(data.weeks) && data.weeks.length) return data.weeks;

  const summary = data.summary || {};
  const goal = Number(summary.weekGoal || 0);

  return [1, 2, 3, 4].map((week) => ({
    name: `Semana ${week}`,
    goal,
    reached: week === 1 ? Number(summary.weekRevenue || 0) : 0,
    missing: week === 1 ? Number(summary.weekMissing || goal) : goal,
    percent: week === 1 ? Number(summary.weekPercent || 0) : 0,
  }));
}

const feriados = [
  "2026-09-07",
  "2026-10-12",
  "2026-11-02",
  "2026-11-20",
  "2026-12-25"
];

function dataChave(data) {
  const ano = data.getFullYear();
  const mes = String(data.getMonth() + 1).padStart(2, "0");
  const dia = String(data.getDate()).padStart(2, "0");

  return `${ano}-${mes}-${dia}`;
}

function contarDiasUteis(inicio, fim) {
  let quantidade = 0;
  const data = new Date(inicio);

  data.setHours(12, 0, 0, 0);

  while (data <= fim) {
    const diaSemana = data.getDay();
    const fimDeSemana = diaSemana === 0 || diaSemana === 6;
    const ehFeriado = feriados.includes(dataChave(data));

    if (!fimDeSemana && !ehFeriado) {
      quantidade++;
    }

    data.setDate(data.getDate() + 1);
  }

  return quantidade;
}

function calcularProjecaoMes(atingido) {
  const hoje = new Date();
  hoje.setHours(12, 0, 0, 0);

  const inicioMes = new Date(
    hoje.getFullYear(),
    hoje.getMonth(),
    1,
    12
  );

  const fimMes = new Date(
    hoje.getFullYear(),
    hoje.getMonth() + 1,
    0,
    12
  );

  const diasUteisDecorridos = contarDiasUteis(inicioMes, hoje);
  const diasUteisMes = contarDiasUteis(inicioMes, fimMes);

  const projecao = diasUteisDecorridos > 0
    ? (Number(atingido || 0) / diasUteisDecorridos) * diasUteisMes
    : 0;

  return {
    projecao,
    diasUteisDecorridos,
    diasUteisMes
  };
}

function renderSummary(data, viewName = activeView) {

  const summary = data.summary || {};

  setText("workbook", data.workbook || "data.json");
  setText("updated", "Atualizado em " + (data.lastModified || "-"));

  const config = views[viewName];

  setText("tableTitle", config.title);
  setText("commitmentLabel", config.commitmentLabel);

  const commitmentValue =
    viewName === "keys" ? summary.weekGoal : summary.commitment;

  if (viewName === "general") {

  setText(
    "commitment",
    formatMoneyCompact(commitmentValue)
  );

  setText(
    "reached",
    formatMoneyCompact(summary.reached)
  );

  const missingValue = Number(summary.missing || 0);

  if (missingValue > 0) {
    setText(
      "missing",
      "-" + formatMoneyCompact(missingValue)
    );
  } else if (missingValue < 0) {
    setText(
      "missing",
      "+" + formatMoneyCompact(Math.abs(missingValue))
    );
  } else {
    setText(
      "missing",
      formatMoneyCompact(0)
    );
  }

} else {

  setText(
    "commitment",
    config.format(commitmentValue)
  );

  setText(
    "reached",
    config.format(summary.reached)
  );

  setText(
    "missing",
    formatBalance(summary.missing, config.format)
  );
}

  byId("missing").className =
    `value ${moneyClass(summary.missing)}`;

  setText("percent", `${summary.percent || 0}%`);

  setText(
    "pendingCount",
    `${summary.pendingCount || 0} ${config.pendingText}`
  );

  setText(
    "positiveCount",
    `${summary.positiveCount || 0} ${config.positiveText}`
  );

  byId("totalProgress").style.width =
    `${Math.min(Number(summary.percent || 0), 100)}%`;

  const projectionCard = byId("projectionCard");
  const cardsContainer = document.querySelector(".cards");

  if (viewName === "general") {

    projectionCard.style.display = "";

    cardsContainer.style.gridTemplateColumns =
  "1.25fr 1.25fr 1.25fr 0.7fr 1.55fr";

    const resultado = calcularProjecaoMes(summary.reached);

    const diferenca =
      resultado.projecao - Number(summary.commitment || 0);

    setText(
      "monthProjection",
      formatMoneyCompact(resultado.projecao)
    );

    setText(
      "projectionDays",
      `${resultado.diasUteisDecorridos} de ${resultado.diasUteisMes} dias úteis`
    );

    if (diferenca >= 0) {

      setText(
        "projectionDifference",
        `+${formatMoney(diferenca)} (acima da meta)`
      );

      byId("projectionDifference").className =
        "projection-difference good";

    } else {

      setText(
        "projectionDifference",
        `-${formatMoney(Math.abs(diferenca))} (abaixo da meta)`
      );

      byId("projectionDifference").className =
        "projection-difference bad";
    }

  } else {

    projectionCard.style.display = "none";

    cardsContainer.style.gridTemplateColumns =
      "repeat(4, minmax(0, 1fr))";
  }
}

    function filteredRows() {
  const term = byId("search").value.trim().toLowerCase();
  const status = byId("status").value;

  const rows = state.rows
  .filter((row) =>
    !term ||
    String(row.seller || "").toLowerCase().includes(term)
  )
  .filter((row) =>
    status === "all" || row.status === status
  )
  .filter((row) => {
    const seller = String(row.seller || "")
      .trim()
      .toUpperCase();

    if (activeView === "general" && seller === "IATAGAM JUNIOR") {
      return false;
    }

    return true;
  });

  // REGRA ESPECIAL SOMENTE PARA O PAINEL GERAL
  if (activeView === "general") {

    // Converte o percentual para número
    const getPercent = (row) => {
      const value = String(row.percent ?? 0)
        .replace("%", "")
        .replace(",", ".");

      return Number(value) || 0;
    };

    // Separa os vendedores comuns
    const varejo = rows.filter((row) => {
      const seller = String(row.seller || "")
        .trim()
        .toUpperCase();

      return !seller.startsWith("KEY");
    });

    // Separa os vendedores KEY
    const keys = rows.filter((row) => {
      const seller = String(row.seller || "")
        .trim()
        .toUpperCase();

      return seller.startsWith("KEY");
    });

    // Varejo: maior % para menor %
    varejo.sort((a, b) =>
      getPercent(b) - getPercent(a)
    );

    // KEY: maior % para menor %
    keys.sort((a, b) =>
      getPercent(b) - getPercent(a)
    );

    // Primeiro Varejo, depois KEY
return [...varejo, ...keys];
}

return rows.sort((a, b) => {
  const getPercent = (row) => {
    const value = String(row.percent ?? 0)
      .replace("%", "")
      .replace(",", ".");

    return Number(value) || 0;
  };

  return getPercent(b) - getPercent(a);
});

} // FECHA A FUNÇÃO filteredRows()

function renderTableHead(config) {
  byId("tableHead").innerHTML = `
        <tr>
          ${config.columns.map((column) => `<th data-sort="${column.key}">${column.label}</th>`).join("")}
          <th>Status</th>
        </tr>
      `;
    }

    function renderTable(viewName = activeView) {
      const config = views[viewName];
      const rows = filteredRows();
      const tableBody = byId("tableBody");
      renderTableHead(config);
      const markup = rows.map((row) => `
        <tr>
          ${config.columns.map((column) => `<td data-label="${column.label}" class="${column.key === "seller" ? "seller" : ""} ${column.className ? column.className(row) : ""}">${column.value(row)}</td>`).join("")}
          <td data-label="Status"><span class="tag ${row.status}">${row.status === "ok" ? "Superou" : "Falta"}</span></td>
        </tr>
      `).join("");
      const duplicateMarkup = rows.map((row) => `
        <tr class="loop-copy">
          ${config.columns.map((column) => `<td data-label="${column.label}" class="${column.key === "seller" ? "seller" : ""} ${column.className ? column.className(row) : ""}">${column.value(row)}</td>`).join("")}
          <td data-label="Status"><span class="tag ${row.status}">${row.status === "ok" ? "Superou" : "Falta"}</span></td>
        </tr>
      `).join("");
      tableBody.innerHTML = rows.length > 8 ? markup + duplicateMarkup : markup;
tableBody.classList.toggle("auto-scroll", rows.length > 8);
requestAnimationFrame(() => startSellerAutoScroll(rows.length > 8));
    }

    function startSellerAutoScroll(enabled) {
  const tableWrap = document.querySelector(".table-wrap");
  const tableBody = byId("tableBody");

  if (sellerScrollFrame) {
    cancelAnimationFrame(sellerScrollFrame);
    sellerScrollFrame = null;
  }

  if (!enabled || window.matchMedia("(max-width: 640px)").matches) {
    sellerScrollOffset = 0;
    tableBody.style.transform = "translateY(0px)";
    return;
  }

  let lastTime = null;
  const speed = 15;

  const originalRows = Array.from(
    tableBody.querySelectorAll("tr:not(.loop-copy)")
  );

  const originalHeight = originalRows.reduce(
    (total, row) => total + row.getBoundingClientRect().height,
    0
  );

  if (originalHeight <= 0) {
    return;
  }

  // Mantém a posição mesmo depois da atualização dos dados
  sellerScrollOffset =
    sellerScrollOffset % originalHeight;

  tableBody.style.transform =
    `translateY(-${sellerScrollOffset}px)`;

  const tick = (time) => {

    if (lastTime === null) {
      lastTime = time;
    }

    const elapsed = time - lastTime;
    lastTime = time;

    sellerScrollOffset +=
      (speed * elapsed) / 1000;

    if (sellerScrollOffset >= originalHeight) {
      sellerScrollOffset = 0;
    }

    tableBody.style.transform =
      `translateY(-${sellerScrollOffset}px)`;

    sellerScrollFrame =
      requestAnimationFrame(tick);
  };

  sellerScrollFrame =
    requestAnimationFrame(tick);
}

    function renderWeeks(viewName = activeView) {
      const config = views[viewName];
      const weeks = state.weeks.slice(0, 5);
      setText("weekCount", `${weeks.length} semanas`);
      const renderWeekCard = (week, loopCopy = false) => {
        const percent = Math.min(Number(week.percent || 0), 100);
        return `
          <div class="week-card${loopCopy ? " loop-copy" : ""}">
            <div class="week-top">
              <div class="week-name">${week.name}</div>
              <div class="week-percent">${Number(week.percent || 0).toFixed(1)}%</div>
            </div>
            <div class="progress"><span style="width:${percent}%"></span></div>
            <div class="week-values">
              <span>Realizado<strong>${config.format(week.reached)}</strong></span>
              <span>Meta<strong>${config.format(week.goal)}</strong></span>
            </div>
          </div>
        `;
      };
      const cards = weeks.map((week) => renderWeekCard(week)).join("");
      const duplicateCards = weeks.map((week) => renderWeekCard(week, true)).join("");
      byId("weekly").innerHTML = weeks.length > 3 ? cards + duplicateCards : cards;
      requestAnimationFrame(() => startWeeklyAutoScroll(weeks.length > 3));
    }

    function startWeeklyAutoScroll(enabled) {
      const weeklyWrap = document.querySelector(".weekly-wrap");
      const weekly = byId("weekly");
      if (weeklyScrollFrame) {
        cancelAnimationFrame(weeklyScrollFrame);
        weeklyScrollFrame = null;
      }
      weekly.style.transform = "translateY(0px)";
      if (!enabled || window.matchMedia("(max-width: 640px)").matches) return;

      const originalCards = Array.from(weekly.querySelectorAll(".week-card:not(.loop-copy)"));
      const loopHeight = originalCards.reduce(
        (total, card) => total + card.getBoundingClientRect().height,
        0
      ) + Math.max(originalCards.length - 1, 0) * 14;
      if (loopHeight <= weeklyWrap.clientHeight) return;

      let lastTime = null;
      const speed = 20;
      let offset = 0;
      const tick = (time) => {
        if (lastTime === null) lastTime = time;
        const elapsed = time - lastTime;
        lastTime = time;
        offset = (offset + (speed * elapsed) / 1000) % loopHeight;
        weekly.style.transform = `translateY(-${offset}px)`;
        weeklyScrollFrame = requestAnimationFrame(tick);
      };
      weeklyScrollFrame = requestAnimationFrame(tick);
    }

    byId("tableHead").addEventListener("click", (event) => {
      const header = event.target.closest("th[data-sort]");
      if (!header) return;
      const key = header.dataset.sort;
      if (state.sortKey === key) {
        state.sortDir = state.sortDir === "asc" ? "desc" : "asc";
      } else {
        state.sortKey = key === "reference" ? "seller" : key;
        state.sortDir = key === "seller" || key === "reference" ? "asc" : "desc";
      }
      renderTable();
    });

    document.querySelectorAll(".view-tab").forEach((button) => {
  button.addEventListener("click", () => {
    activeView = button.dataset.view;
    state.sortKey = "missing";
    state.sortDir = "desc";

    document.querySelectorAll(".view-tab").forEach((tab) => {
      tab.classList.toggle("active", tab === button);
    });

    loadData();
  });
});

byId("search").addEventListener("input", renderTable);
byId("status").addEventListener("change", renderTable);
byId("refresh").addEventListener("click", loadData);

const autoRotateViews = ["sales", "general"];
let autoRotateTimer = null;

function changeToNextView() {
  const currentIndex = autoRotateViews.indexOf(activeView);
  const nextIndex = (currentIndex + 1) % autoRotateViews.length;

  activeView = autoRotateViews[nextIndex];

  state.sortKey = "missing";
  state.sortDir = "desc";

  document.querySelectorAll(".view-tab").forEach((tab) => {
    tab.classList.toggle("active", tab.dataset.view === activeView);
  });

  loadData();
}

function startAutoRotate() {
  if (autoRotateTimer) clearInterval(autoRotateTimer);
  autoRotateTimer = setInterval(changeToNextView, 120000);
}

loadData();
setInterval(loadData, 30000);
startAutoRotate();
  </script>
</body>
</html>
"""


def get_local_addresses():
    addresses = {"127.0.0.1", "localhost"}
    try:
        hostname = socket.gethostname()
        addresses.add(hostname)
        for info in socket.getaddrinfo(hostname, None, family=socket.AF_INET):
            ip = info[4][0]
            if ip and not ip.startswith("127."):
                addresses.add(ip)
    except OSError:
        pass
    return [f"http://{address}:{PORT}" for address in sorted(addresses)]


def as_number(value):
    if value is None or value == "":
        return 0.0
    if isinstance(value, str) and value.strip().startswith("#"):
        raise ValueError(f"Formula com erro no Excel: {value}")
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def money(value):
    return round(as_number(value), 2)


def find_compromisso_workbook():
    if PYTHON_EXCEL_PATH.exists():
        return PYTHON_EXCEL_PATH
    if EXCEL_PATH.exists():
        return EXCEL_PATH
    for pattern in (
        "BASE PYTHON*.xlsx",
        "BASE PYTHON*.xlsm",
        "COMPROMISSO MES*.xlsx",
        "COMPROMISSO MES*.xlsm",
        "COMPROMISSO MÊS*.xlsx",
        "COMPROMISSO MÊS*.xlsm",
        "COMPROMISSO MAIO*.xlsx",
        "COMPROMISSO MAIO*.xlsm",
    ):
        candidates = [
            path
            for path in BASE_DIR.glob(pattern)
            if not path.name.startswith("~$")
        ]
        if candidates:
            return max(candidates, key=lambda path: path.stat().st_mtime)
    return None


def default_weekly_goals():
    return {f"Semana {index}": DEFAULT_WEEKLY_GOAL for index in range(1, 6)}


def read_weekly_goals():
    goals = default_weekly_goals()
    if WEEKLY_GOALS_PATH.exists():
        try:
            stored_goals = json.loads(WEEKLY_GOALS_PATH.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            stored_goals = {}
        for name, value in stored_goals.items():
            goals[normalize_week_name(name)] = money(value)
    return goals


def save_weekly_goals(goals):
    normalized_goals = default_weekly_goals()
    for name, value in goals.items():
        normalized_goals[normalize_week_name(name)] = money(value)
    WEEKLY_GOALS_PATH.write_text(
        json.dumps(normalized_goals, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return normalized_goals


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_key(value):
    text = normalize_text(value).upper()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return " ".join(text.split())


def find_header_row(sheet, required_headers, max_rows=25):
    required = {normalize_key(header) for header in required_headers}
    for row_number in range(1, min(sheet.max_row, max_rows) + 1):
        headers = {
            normalize_key(sheet.cell(row_number, column).value): column
            for column in range(1, sheet.max_column + 1)
            if normalize_key(sheet.cell(row_number, column).value)
        }
        if required.issubset(headers):
            return row_number, headers
    raise ValueError(f"Cabecalhos nao encontrados em {sheet.title}: {', '.join(required_headers)}")


def normalize_week_name(value):
    text = normalize_text(value).title()
    digits = "".join(ch for ch in text if ch.isdigit())
    if digits:
        return f"Semana {int(digits)}"
    return text or "Semana 1"


def week_sort_key(item):
    digits = "".join(ch for ch in item["name"] if ch.isdigit())
    return int(digits) if digits else 99


def apply_weekly_goals(data):
    goals = read_excel_weekly_goals() or read_weekly_goals()
    weeks = data.get("weeks") or []
    for index in range(1, 6):
        week_name = f"Semana {index}"
        if not any(normalize_week_name(item.get("name")) == week_name for item in weeks):
            weeks.append({"name": week_name, "goal": 0, "reached": 0, "missing": 0, "percent": 0})

    for week in weeks:
        week_name = normalize_week_name(week.get("name"))
        week["name"] = week_name
        week["goal"] = money(goals.get(week_name, DEFAULT_WEEKLY_GOAL))
        week["reached"] = money(week.get("reached"))
        week["missing"] = money(week["goal"] - week["reached"])
        week["percent"] = round((week["reached"] / week["goal"] * 100) if week["goal"] else 0, 1)

    data["weeks"] = sorted(weeks, key=week_sort_key)[:5]
    summary = data.setdefault("summary", {})
    current_week = data["weeks"][0] if data["weeks"] else {}
    summary["weekGoal"] = current_week.get("goal", 0)
    summary["weekRevenue"] = current_week.get("reached", 0)
    summary["weekMissing"] = current_week.get("missing", 0)
    summary["weekPercent"] = current_week.get("percent", 0)
    return data


def read_sales_rows(workbook):
    sheet = workbook["Planilha2"] if "Planilha2" in workbook.sheetnames else workbook["BASE"]
    header_row, headers = find_header_row(sheet, ["VENDEDOR", "COMPROMISSO", "ATINGIDO"])
    seller_col = headers.get("NOME", headers["VENDEDOR"])
    commitment_col = headers["COMPROMISSO"]
    reached_col = headers["ATINGIDO"]
    missing_col = headers.get("FALTA")
    average_col = headers.get("MEDIA")
    difference_col = headers.get("DIFERENCA")
    rows = []
    for row_number in range(header_row + 1, sheet.max_row + 1):
        seller = normalize_text(sheet.cell(row_number, seller_col).value)
        if not seller:
            continue
        if seller.upper() == "TOTAL":
            break
        if seller.upper() == "VENDEDOR":
            continue

        commitment = money(sheet.cell(row_number, commitment_col).value)
        reached = money(sheet.cell(row_number, reached_col).value)
        missing = money(sheet.cell(row_number, missing_col).value) if missing_col else money(commitment - reached)
        average = money(sheet.cell(row_number, average_col).value) if average_col else 0.0
        difference = money(sheet.cell(row_number, difference_col).value) if difference_col else 0.0
        percent = (reached / commitment * 100) if commitment else 0

        rows.append(
            {
                "seller": seller,
                "commitment": commitment,
                "reached": reached,
                "missing": missing,
                "average": average,
                "difference": difference,
                "percent": round(percent, 1),
                "status": "ok" if missing <= 0 else "pending",
            }
        )
    return rows


def read_excel_weekly_goals():
    excel_path = find_compromisso_workbook()
    if not excel_path:
        return {}

    workbook = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    try:
        if excel_path.name.upper().startswith("BASE PYTHON") and "BASE" in workbook.sheetnames:
            data = read_base_python_dashboard_data(excel_path, workbook)
        elif excel_path.name.upper().startswith("BASE PYTHON") and "Planilha1" in workbook.sheetnames:
            data = read_base_python_planilha1_data(excel_path, workbook)
        elif "Planilha1" in workbook.sheetnames:
            weeks, _daily_totals = read_weekly_blocks(workbook)
            data = {"weeks": weeks}
        else:
            return {}
    finally:
        workbook.close()

    return {
        normalize_week_name(week.get("name")): money(week.get("goal"))
        for week in data.get("weeks", [])
        if money(week.get("goal")) > 0
    }


def read_month_summary_data(workbook):
    sheet = workbook["RESUMO"]
    indicators = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        key = normalize_key(row[0] if row else None)
        if key:
            indicators[key] = money(row[1] if len(row) > 1 else 0)

    weeks = []
    for index in range(1, 6):
        reached = money(indicators.get(f"SEMANA {index}", 0))
        weeks.append(
            {
                "name": f"Semana {index}",
                "goal": DEFAULT_WEEKLY_GOAL,
                "reached": reached,
                "missing": money(DEFAULT_WEEKLY_GOAL - reached),
                "percent": round((reached / DEFAULT_WEEKLY_GOAL * 100) if DEFAULT_WEEKLY_GOAL else 0, 1),
            }
        )

    return {
        "commitment": money(indicators.get("META DIA", 0)),
        "reached": money(indicators.get("ATINGIDO", 0)),
        "missing": money(indicators.get("FALTA", 0)),
        "weeks": weeks,
    }


def read_weekly_blocks(workbook):
    sheet = workbook["Planilha1"]
    weeks = []
    daily_totals = []
    day_groups = [
        (3, 4, 5),
        (6, 7, 8),
        (9, 10, 11),
        (12, 13, 14),
        (15, 16, 17),
    ]

    for row_number in range(1, sheet.max_row + 1):
        week_name = normalize_text(sheet.cell(row_number, 2).value)
        if not week_name.upper().startswith("SEMANA"):
            continue

        total_row = None
        for candidate in range(row_number + 2, min(sheet.max_row, row_number + 30) + 1):
            if normalize_text(sheet.cell(candidate, 2).value).upper() == "TOTAL":
                total_row = candidate
                break
        if not total_row:
            continue

        week_goal = 0.0
        week_reached = 0.0
        for commitment_col, reached_col, missing_col in day_groups:
            date_value = sheet.cell(row_number, commitment_col + 1).value
            day_name = normalize_text(sheet.cell(row_number, commitment_col).value)
            commitment = money(sheet.cell(total_row, commitment_col).value)
            reached = money(sheet.cell(total_row, reached_col).value)
            missing = money(sheet.cell(total_row, missing_col).value)

            has_daily_value = commitment or reached or missing
            if not has_daily_value:
                continue

            week_goal = money(week_goal + commitment)
            week_reached = money(week_reached + reached)
            if isinstance(date_value, datetime):
                daily_date = date_value.date()
            else:
                daily_date = None
            daily_totals.append(
                {
                    "week": week_name,
                    "day": day_name,
                    "date": daily_date,
                    "commitment": commitment,
                    "reached": reached,
                    "missing": missing,
                }
            )

        week_missing = money(week_goal - week_reached)
        week_percent = (week_reached / week_goal * 100) if week_goal else 0
        weeks.append(
            {
                "name": week_name.title(),
                "goal": week_goal,
                "reached": week_reached,
                "missing": week_missing,
                "percent": round(week_percent, 1),
            }
        )

    return sorted(weeks, key=week_sort_key)[:5], daily_totals


def read_base_python_dashboard_data(excel_path, workbook):
    sheet = workbook["BASE"]
    header_row = 9
    seller_col = 8
    commitment_col = 9
    reached_col = 10
    missing_col = 11

    rows = []
    for row_number in range(header_row + 1, sheet.max_row + 1):
        seller = normalize_text(sheet.cell(row_number, seller_col).value)
        if not seller:
            continue
        if normalize_key(seller) in {"SEMANAL", "TOTAL", "VENDEDOR"}:
            break

        commitment = money(sheet.cell(row_number, commitment_col).value)
        reached = money(sheet.cell(row_number, reached_col).value)
        missing = money(sheet.cell(row_number, missing_col).value) if missing_col else money(commitment - reached)
        percent = (reached / commitment * 100) if commitment else 0
        rows.append(
            {
                "seller": seller,
                "commitment": commitment,
                "reached": reached,
                "missing": missing,
                "average": 0.0,
                "difference": 0.0,
                "percent": round(percent, 1),
                "status": "ok" if missing <= 0 else "pending",
            }
        )

    indicators = {}
    for row_number in range(1, min(sheet.max_row, 8) + 1):
        key = normalize_key(sheet.cell(row_number, 8).value)
        if key:
            indicators[key] = money(sheet.cell(row_number, 9).value)

    weeks = []
    for row_number in range(1, sheet.max_row + 1):
        raw_label = normalize_text(sheet.cell(row_number, seller_col).value)
        if not normalize_key(raw_label).startswith("SEMANA"):
            continue
        label = normalize_week_name(raw_label)
        goal = round(money(sheet.cell(row_number, commitment_col).value))
        reached = money(sheet.cell(row_number, reached_col).value)
        missing = money(sheet.cell(row_number, missing_col).value)
        weeks.append(
            {
                "name": label,
                "goal": goal,
                "reached": reached,
                "missing": missing,
                "percent": round((reached / goal * 100) if goal else 0, 1),
            }
        )

    total_commitment = money(indicators.get("META DIA", sum(item["commitment"] for item in rows)))
    total_reached = money(indicators.get("ATINGIDO", sum(item["reached"] for item in rows)))
    total_missing = money(indicators.get("FALTA", total_commitment - total_reached))
    total_percent = (total_reached / total_commitment * 100) if total_commitment else 0

    return {
        "workbook": excel_path.name,
        "lastModified": datetime.fromtimestamp(excel_path.stat().st_mtime).strftime(
            "%d/%m/%Y %H:%M"
        ),
        "summary": {
            "commitment": total_commitment,
            "reached": total_reached,
            "missing": total_missing,
            "percent": round(total_percent, 1),
            "weekGoal": weeks[0]["goal"] if weeks else 0,
            "weekRevenue": weeks[0]["reached"] if weeks else 0,
            "weekMissing": weeks[0]["missing"] if weeks else 0,
            "weekPercent": weeks[0]["percent"] if weeks else 0,
            "positiveCount": sum(1 for item in rows if item["missing"] <= 0),
            "pendingCount": sum(1 for item in rows if item["missing"] > 0),
        },
        "rows": sorted(rows, key=lambda item: item["missing"], reverse=True),
        "history": [],
        "weeks": weeks,
    }


def find_block_start(sheet, title):
    wanted = normalize_key(title)
    for column in range(1, sheet.max_column + 1):
        if normalize_key(sheet.cell(1, column).value) == wanted:
            return column
    raise ValueError(f"Bloco nao encontrado em {sheet.title}: {title}")


def find_block_start_any(sheet, titles):
    for title in titles:
        try:
            return find_block_start(sheet, title)
        except ValueError:
            continue
    raise ValueError(f"Bloco nao encontrado em {sheet.title}: {', '.join(titles)}")

def read_base_python_planilha1_data(excel_path, workbook):
    sheet = workbook["Planilha1"]
    block_start = find_block_start(sheet, "VAREJO")

    seller_col = block_start - 1
    reference_col = block_start
    commitment_col = block_start + 1
    reached_col = block_start + 2
    missing_col = block_start + 3

    rows = []

    for row_number in range(10, sheet.max_row + 1):
        seller = normalize_text(
            sheet.cell(row_number, seller_col).value
        )

        reference = normalize_text(
            sheet.cell(row_number, reference_col).value
        )

        if not seller:
            continue

        if normalize_key(seller) in {
            "SEMANAL",
            "TOTAL",
            "VENDEDOR"
        }:
            break

        commitment = money(
            sheet.cell(row_number, commitment_col).value
        )

        reached = money(
            sheet.cell(row_number, reached_col).value
        )

        missing = money(
            sheet.cell(row_number, missing_col).value
        )

        percent = (
            reached / commitment * 100
            if commitment
            else 0
        )

        rows.append(
            {
                "seller": seller,
                "reference": reference or seller,
                "commitment": commitment,
                "reached": reached,
                "missing": missing,
                "average": 0.0,
                "difference": 0.0,
                "percent": round(percent, 1),
                "status": "ok" if missing <= 0 else "pending",
            }
        )

    indicators = {}

    for row_number in range(4, 8):
        key = normalize_key(
            sheet.cell(row_number, reference_col).value
        )

        if key:
            raw_value = sheet.cell(
                row_number,
                commitment_col
            ).value

            if (
                isinstance(raw_value, str)
                and raw_value.strip().startswith("#")
            ):
                continue

            indicators[key] = (
                round(as_number(raw_value), 4)
                if key == "REALIZADO"
                else money(raw_value)
            )

    weeks = []
    weekly_header_row = None

    for row_number in range(10, sheet.max_row + 1):
        if (
            normalize_key(
                sheet.cell(row_number, reference_col).value
            )
            == "SEMANAL"
        ):
            weekly_header_row = row_number
            break

    if weekly_header_row:
        for row_number in range(
            weekly_header_row + 1,
            sheet.max_row + 1
        ):
            raw_label = normalize_text(
                sheet.cell(row_number, reference_col).value
            )

            if not normalize_key(raw_label).startswith("SEMANA"):
                continue

            goal = money(
                sheet.cell(row_number, commitment_col).value
            )

            reached = money(
                sheet.cell(row_number, reached_col).value
            )

            missing = money(
                sheet.cell(row_number, missing_col).value
            )

            weeks.append(
                {
                    "name": normalize_week_name(raw_label),
                    "goal": goal,
                    "reached": reached,
                    "missing": missing,
                    "percent": round(
                        (reached / goal * 100)
                        if goal
                        else 0,
                        1
                    ),
                }
            )

    total_commitment = money(
        indicators.get(
            "META DIA",
            sum(item["commitment"] for item in rows)
        )
    )

    total_reached = money(
        indicators.get(
            "ATINGIDO",
            sum(item["reached"] for item in rows)
        )
    )

    total_missing = money(
        indicators.get(
            "FALTA",
            total_commitment - total_reached
        )
    )

    total_percent = (
        total_reached / total_commitment * 100
        if total_commitment
        else 0
    )

    return {
        "workbook": excel_path.name,

        "lastModified": datetime.fromtimestamp(
            excel_path.stat().st_mtime
        ).strftime("%d/%m/%Y %H:%M"),

        "summary": {
            "commitment": total_commitment,
            "reached": total_reached,
            "missing": total_missing,
            "percent": round(total_percent, 1),

            "weekGoal": weeks[0]["goal"] if weeks else 0,
            "weekRevenue": weeks[0]["reached"] if weeks else 0,
            "weekMissing": weeks[0]["missing"] if weeks else 0,
            "weekPercent": weeks[0]["percent"] if weeks else 0,

            "positiveCount": sum(
                1 for item in rows
                if item["missing"] <= 0
            ),

            "pendingCount": sum(
                1 for item in rows
                if item["missing"] > 0
            ),
        },

        "rows": sorted(
            rows,
            key=lambda item: item["missing"],
            reverse=True
        ),

        "history": [],
        "weeks": weeks,
    }

def find_block_table_header(
    sheet,
    reference_col,
    seller_col,
    commitment_col,
    reached_col,
    max_rows=25
):
    for row_number in range(
        1,
        min(sheet.max_row, max_rows) + 1
    ):
        reference = normalize_key(
            sheet.cell(row_number, reference_col).value
        )

        seller = normalize_key(
            sheet.cell(row_number, seller_col).value
        )

        commitment = normalize_key(
            sheet.cell(row_number, commitment_col).value
        )

        reached = normalize_key(
            sheet.cell(row_number, reached_col).value
        )

        if (
            reference in {"REFERENCIA", "REFERENCA"}
            and seller == "VENDEDOR"
            and commitment in {"COMPROMISSO", "META", "META A"}
            and reached == "ATINGIDO"
        ):
            return row_number

    raise ValueError("Cabecalho do bloco nao encontrado")

def read_general_planilha1_data(excel_path, workbook):
    sheet = workbook["Planilha1"]
    title_col = find_block_start(sheet, "GERAL")

    reference_col = title_col - 1
    seller_col = title_col
    commitment_col = title_col + 1
    reached_col = title_col + 2
    missing_col = title_col + 3

    header_row = find_block_table_header(
        sheet,
        reference_col,
        seller_col,
        commitment_col,
        reached_col
    )

    rows = []

    for row_number in range(header_row + 1, sheet.max_row + 1):
        reference = normalize_text(
            sheet.cell(row_number, reference_col).value
        )

        seller = normalize_text(
            sheet.cell(row_number, seller_col).value
        )

        if not reference and not seller:
            break

        if (
            normalize_key(reference) in {"SEMANAL", "TOTAL"}
            or normalize_key(seller) in {"SEMANAL", "TOTAL"}
        ):
            break

        commitment = money(
            sheet.cell(row_number, commitment_col).value
        )

        reached = money(
            sheet.cell(row_number, reached_col).value
        )

        missing = money(
            sheet.cell(row_number, missing_col).value
        )

        percent = (
            reached / commitment * 100
            if commitment
            else 0
        )

        rows.append(
            {
                "reference": reference or seller,
                "seller": seller or reference,
                "commitment": commitment,
                "reached": reached,
                "missing": missing,
                "average": 0.0,
                "difference": 0.0,
                "percent": round(percent, 1),
                "status": "ok" if missing <= 0 else "pending",
            }
        )

    indicators = {}

    for row_number in range(4, 8):
        key = normalize_key(
            sheet.cell(row_number, title_col).value
        )

        if key:
            raw_value = sheet.cell(
                row_number,
                commitment_col
            ).value

            if (
                isinstance(raw_value, str)
                and raw_value.strip().startswith("#")
            ):
                continue

            indicators[key] = (
                round(as_number(raw_value), 4)
                if key == "REALIZADO"
                else money(raw_value)
            )

    weeks = []
    weekly_header_row = None

    for row_number in range(
        header_row + 1,
        sheet.max_row + 1
    ):
        if (
            normalize_key(
                sheet.cell(row_number, reference_col).value
            )
            == "SEMANAL"
        ):
            weekly_header_row = row_number
            break

    if weekly_header_row:
        for row_number in range(
            weekly_header_row + 1,
            sheet.max_row + 1
        ):
            raw_label = normalize_text(
                sheet.cell(row_number, reference_col).value
            )

            if not normalize_key(raw_label).startswith("SEMANA"):
                continue

            goal = money(
                sheet.cell(row_number, commitment_col).value
            )

            reached = money(
                sheet.cell(row_number, reached_col).value
            )

            missing = money(
                sheet.cell(row_number, missing_col).value
            )

            weeks.append(
                {
                    "name": normalize_week_name(raw_label),
                    "goal": goal,
                    "reached": reached,
                    "missing": missing,
                    "percent": round(
                        (reached / goal * 100)
                        if goal
                        else 0,
                        1
                    ),
                }
            )

    total_commitment = money(
        indicators.get(
            "META MES",
            sum(item["commitment"] for item in rows)
        )
    )

    total_reached = money(
        indicators.get(
            "ATINGIDO",
            sum(item["reached"] for item in rows)
        )
    )

    total_missing = money(
        indicators.get(
            "FALTA",
            total_commitment - total_reached
        )
    )

    realized_indicator = indicators.get("REALIZADO")

    total_percent = (
        realized_indicator * 100
        if realized_indicator is not None
        else (
            total_reached / total_commitment * 100
            if total_commitment
            else 0
        )
    )

    return {
        "workbook": excel_path.name,

        "lastModified": datetime.fromtimestamp(
            excel_path.stat().st_mtime
        ).strftime("%d/%m/%Y %H:%M"),

        "summary": {
            "commitment": total_commitment,
            "reached": total_reached,
            "missing": total_missing,
            "percent": round(total_percent, 1),

            "weekGoal": weeks[0]["goal"] if weeks else 0,
            "weekRevenue": weeks[0]["reached"] if weeks else 0,
            "weekMissing": weeks[0]["missing"] if weeks else 0,
            "weekPercent": weeks[0]["percent"] if weeks else 0,

            "positiveCount": sum(
                1
                for item in rows
                if item["missing"] <= 0
            ),

            "pendingCount": sum(
                1
                for item in rows
                if item["missing"] > 0
            ),
        },

        "rows": sorted(
            rows,
            key=lambda item: item["missing"],
            reverse=True
        ),

        "history": [],
        "weeks": weeks,
    }


def read_positivacao_milho_planilha1_data(excel_path, workbook):
    sheet = workbook["Planilha1"]
    title_col = find_block_start_any(
        sheet,
        ("POSITIVACAO MILHO", "POSITIVACAO BRIOCHE")
    )

    seller_col = title_col - 1
    reference_col = title_col
    commitment_col = title_col + 1
    reached_col = title_col + 2
    missing_col = title_col + 3

    rows = []

    for row_number in range(10, sheet.max_row + 1):
        reference = normalize_text(
            sheet.cell(row_number, reference_col).value
        )

        seller = normalize_text(
            sheet.cell(row_number, seller_col).value
        )

        if not reference and not seller:
            continue

        if (
            normalize_key(reference)
            in {"SEMANAL", "TOTAL", "REFERENCA"}
            or normalize_key(seller)
            in {"SEMANAL", "TOTAL", "VENDEDOR"}
        ):
            break

        commitment = money(
            sheet.cell(row_number, commitment_col).value
        )

        reached = money(
            sheet.cell(row_number, reached_col).value
        )

        missing = money(
            sheet.cell(row_number, missing_col).value
        )

        percent = (
            reached / commitment * 100
            if commitment
            else 0
        )

        rows.append(
            {
                "reference": reference or seller,
                "seller": seller or reference,
                "commitment": commitment,
                "reached": reached,
                "missing": missing,
                "average": 0.0,
                "difference": 0.0,
                "percent": round(percent, 1),
                "status": "ok" if missing <= 0 else "pending",
            }
        )

    indicators = {}

    for row_number in range(4, 8):
        key = normalize_key(
            sheet.cell(row_number, seller_col).value
        )

        if key:
            raw_value = sheet.cell(
                row_number,
                commitment_col
            ).value

            indicators[key] = (
                round(as_number(raw_value), 4)
                if key == "REALIZADO"
                else money(raw_value)
            )

    weeks = []
    weekly_header_row = None

    for row_number in range(10, sheet.max_row + 1):
        if (
            normalize_key(
                sheet.cell(row_number, reference_col).value
            )
            == "SEMANAL"
        ):
            weekly_header_row = row_number
            break

    if weekly_header_row:
        for row_number in range(
            weekly_header_row + 1,
            sheet.max_row + 1
        ):
            raw_label = normalize_text(
                sheet.cell(row_number, reference_col).value
            )

            if not normalize_key(raw_label).startswith("SEMANA"):
                continue

            goal = POSITIVACAO_MILHO_WEEKLY_GOAL

            reached = money(
                sheet.cell(row_number, reached_col).value
            )

            missing = money(goal - reached)

            weeks.append(
                {
                    "name": normalize_week_name(raw_label),
                    "goal": goal,
                    "reached": reached,
                    "missing": missing,
                    "percent": round(
                        (reached / goal * 100)
                        if goal
                        else 0,
                        1
                    ),
                }
            )

    total_commitment = money(
        indicators.get(
            "META MES",
            sum(item["commitment"] for item in rows)
        )
    )

    total_reached = money(
        indicators.get(
            "ATINGIDO",
            sum(item["reached"] for item in rows)
        )
    )

    total_missing = money(
        indicators.get(
            "FALTA",
            total_commitment - total_reached
        )
    )

    realized_indicator = indicators.get("REALIZADO")

    total_percent = (
        realized_indicator * 100
        if realized_indicator is not None
        else (
            total_reached / total_commitment * 100
            if total_commitment
            else 0
        )
    )

    return {
        "workbook": excel_path.name,

        "lastModified": datetime.fromtimestamp(
            excel_path.stat().st_mtime
        ).strftime("%d/%m/%Y %H:%M"),

        "summary": {
            "commitment": total_commitment,
            "reached": total_reached,
            "missing": total_missing,
            "percent": round(total_percent, 1),

            "weekGoal": weeks[0]["goal"] if weeks else 0,
            "weekRevenue": weeks[0]["reached"] if weeks else 0,
            "weekMissing": weeks[0]["missing"] if weeks else 0,
            "weekPercent": weeks[0]["percent"] if weeks else 0,

            "positiveCount": sum(
                1 for item in rows
                if item["missing"] <= 0
            ),

            "pendingCount": sum(
                1 for item in rows
                if item["missing"] > 0
            ),
        },

        "rows": sorted(
            rows,
            key=lambda item: item["missing"],
            reverse=True
        ),

        "history": [],
        "weeks": weeks,
    }

def read_keys_planilha1_data(excel_path, workbook):
    sheet = workbook["Planilha1"]

    title_col = find_block_start(sheet, "KEY")
    reference_col = title_col - 1
    seller_col = title_col
    commitment_col = title_col + 1
    reached_col = title_col + 2
    missing_col = title_col + 3

    header_row = find_block_table_header(
        sheet,
        reference_col,
        seller_col,
        commitment_col,
        reached_col
    )

    rows = []

    for row_number in range(
        header_row + 1,
        sheet.max_row + 1
    ):
        reference = normalize_text(
            sheet.cell(row_number, reference_col).value
        )

        seller = normalize_text(
            sheet.cell(row_number, seller_col).value
        )

        if not reference and not seller:
            break

        if (
            normalize_key(reference) in {"SEMANAL", "TOTAL"}
            or normalize_key(seller) in {"SEMANAL", "TOTAL"}
        ):
            break

        if normalize_key(seller) == "KEY PIRACICABA":
            continue

        commitment = money(
            sheet.cell(row_number, commitment_col).value
        )

        reached = money(
            sheet.cell(row_number, reached_col).value
        )

        missing = money(
            sheet.cell(row_number, missing_col).value
        )

        percent = (
            reached / commitment * 100
            if commitment
            else 0
        )

        rows.append(
            {
                "reference": reference or seller,
                "seller": seller or reference,
                "commitment": commitment,
                "reached": reached,
                "missing": missing,
                "average": 0.0,
                "difference": 0.0,
                "percent": round(percent, 1),
                "status": "ok" if missing <= 0 else "pending",
            }
        )

    indicators = {}

    for row_number in range(4, 8):

        reference_key = normalize_key(
            sheet.cell(row_number, reference_col).value
        )

        seller_key = normalize_key(
            sheet.cell(row_number, seller_col).value
        )

        key = seller_key or reference_key

        if key:
            raw_value = sheet.cell(
                row_number,
                commitment_col
            ).value

            if (
                isinstance(raw_value, str)
                and raw_value.strip().startswith("#")
            ):
                continue

            indicators[key] = (
                round(as_number(raw_value), 4)
                if key == "REALIZADO"
                else money(raw_value)
            )

    weeks = []
    weekly_header_row = None
    weekly_label_col = None

    for row_number in range(
        header_row + 1,
        sheet.max_row + 1
    ):
        reference_value = normalize_key(
            sheet.cell(row_number, reference_col).value
        )

        seller_value = normalize_key(
            sheet.cell(row_number, seller_col).value
        )

        if seller_value == "SEMANAL":
            weekly_header_row = row_number
            weekly_label_col = seller_col
            break

        if reference_value == "SEMANAL":
            weekly_header_row = row_number
            weekly_label_col = reference_col
            break

    if weekly_header_row:
        for row_number in range(
            weekly_header_row + 1,
            sheet.max_row + 1
        ):
            raw_label = normalize_text(
                sheet.cell(row_number, weekly_label_col).value
            )

            if not raw_label:
                break

            if not normalize_key(raw_label).startswith("SEMANA"):
                continue

            goal = money(
                sheet.cell(row_number, commitment_col).value
            )

            reached = money(
                sheet.cell(row_number, reached_col).value
            )

            missing = money(
                sheet.cell(row_number, missing_col).value
            )

            weeks.append(
                {
                    "name": normalize_week_name(raw_label),
                    "goal": goal,
                    "reached": reached,
                    "missing": missing,
                    "percent": round(
                        (reached / goal * 100)
                        if goal
                        else 0,
                        1
                    ),
                }
            )

    total_commitment = money(
        indicators.get(
            "META SEMANAL",
            sum(item["commitment"] for item in rows)
        )
    )

    total_reached = money(
        indicators.get(
            "ATINGIDO",
            sum(item["reached"] for item in rows)
        )
    )

    total_missing = money(
        indicators.get(
            "FALTA",
            total_commitment - total_reached
        )
    )

    realized_indicator = indicators.get("REALIZADO")

    total_percent = (
        realized_indicator * 100
        if realized_indicator is not None
        else (
            total_reached / total_commitment * 100
            if total_commitment
            else 0
        )
    )

    return {
        "workbook": excel_path.name,
        "lastModified": datetime.fromtimestamp(
            excel_path.stat().st_mtime
        ).strftime("%d/%m/%Y %H:%M"),

        "summary": {
            "commitment": total_commitment,
            "reached": total_reached,
            "missing": total_missing,
            "percent": round(total_percent, 1),

            "weekGoal": weeks[0]["goal"] if weeks else 0,
            "weekRevenue": weeks[0]["reached"] if weeks else 0,
            "weekMissing": weeks[0]["missing"] if weeks else 0,
            "weekPercent": weeks[0]["percent"] if weeks else 0,

            "positiveCount": sum(
                1 for item in rows
                if item["missing"] <= 0
            ),

            "pendingCount": sum(
                1 for item in rows
                if item["missing"] > 0
            ),
        },

        "rows": sorted(
            rows,
            key=lambda item: item["missing"],
            reverse=True
        ),

        "history": [],
        "weeks": weeks,
    }

def select_daily_total(daily_totals):
    if not daily_totals:
        return None

    today = datetime.now().date()
    for item in daily_totals:
        if item["date"] == today:
            return item

    dated_items = [item for item in daily_totals if item["date"] and item["date"] <= today]
    if dated_items:
        return max(dated_items, key=lambda item: item["date"])

    return daily_totals[-1]


def read_compromisso_maio_data():
    excel_path = find_compromisso_workbook()
    if not excel_path:
        raise FileNotFoundError(f"Arquivo nao encontrado: {EXCEL_PATH.name}")

    workbook = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    try:
        if excel_path.name.upper().startswith("BASE PYTHON") and "BASE" in workbook.sheetnames:
            return read_base_python_dashboard_data(excel_path, workbook)
        if excel_path.name.upper().startswith("BASE PYTHON") and "Planilha1" in workbook.sheetnames:
            return read_base_python_planilha1_data(excel_path, workbook)

        rows = read_sales_rows(workbook)
        if "RESUMO" in workbook.sheetnames and "BASE" in workbook.sheetnames:
            month_summary = read_month_summary_data(workbook)
            weeks = month_summary["weeks"]
            daily_totals = []
        else:
            month_summary = None
            weeks, daily_totals = read_weekly_blocks(workbook)
    finally:
        workbook.close()

    daily_total = select_daily_total(daily_totals)
    total_commitment = money(
        month_summary["commitment"]
        if month_summary
        else daily_total["commitment"]
        if daily_total
        else sum(item["commitment"] for item in rows)
    )
    total_reached = money(
        month_summary["reached"]
        if month_summary
        else daily_total["reached"]
        if daily_total
        else sum(item["reached"] for item in rows)
    )
    total_missing = money(
        month_summary["missing"]
        if month_summary
        else daily_total["missing"]
        if daily_total
        else total_commitment - total_reached
    )
    total_percent = (total_reached / total_commitment * 100) if total_commitment else 0

    return {
        "workbook": excel_path.name,
        "lastModified": datetime.fromtimestamp(excel_path.stat().st_mtime).strftime(
            "%d/%m/%Y %H:%M"
        ),
        "summary": {
            "commitment": total_commitment,
            "reached": total_reached,
            "missing": total_missing,
            "percent": round(total_percent, 1),
            "weekGoal": weeks[0]["goal"] if weeks else 0,
            "weekRevenue": weeks[0]["reached"] if weeks else 0,
            "weekMissing": weeks[0]["missing"] if weeks else 0,
            "weekPercent": weeks[0]["percent"] if weeks else 0,
            "positiveCount": sum(1 for item in rows if item["missing"] <= 0),
            "pendingCount": sum(1 for item in rows if item["missing"] > 0),
        },
        "rows": sorted(rows, key=lambda item: item["missing"], reverse=True),
        "history": [],
        "weeks": weeks,
    }


def read_legacy_excel_dashboard_data():
    workbook = openpyxl.load_workbook(LEGACY_EXCEL_PATH, data_only=True, read_only=True)
    try:
        sheet = workbook["Vendas"]
        rows_by_seller = {}
        weeks_by_name = {}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            date_value, week, seller, goal, reached, *_ = list(row) + [None] * 2
            if not seller:
                continue

            seller_name = str(seller).strip()
            week_name = str(week or "Semana 1").strip()
            goal_value = money(goal)
            reached_value = money(reached)

            seller_row = rows_by_seller.setdefault(
                seller_name,
                {"seller": seller_name, "commitment": 0.0, "reached": 0.0},
            )
            seller_row["commitment"] = money(seller_row["commitment"] + goal_value)
            seller_row["reached"] = money(seller_row["reached"] + reached_value)

            week_row = weeks_by_name.setdefault(
                week_name,
                {"name": week_name, "goal": 0.0, "reached": 0.0},
            )
            week_row["goal"] = money(week_row["goal"] + goal_value)
            week_row["reached"] = money(week_row["reached"] + reached_value)
    finally:
        workbook.close()

    rows = []
    for item in rows_by_seller.values():
        missing = money(item["commitment"] - item["reached"])
        percent = (item["reached"] / item["commitment"] * 100) if item["commitment"] else 0
        rows.append(
            {
                **item,
                "missing": missing,
                "average": 0.0,
                "difference": 0.0,
                "percent": round(percent, 1),
                "status": "ok" if missing <= 0 else "pending",
            }
        )

    weeks = []
    for item in sorted(weeks_by_name.values(), key=week_sort_key)[:5]:
        missing = money(item["goal"] - item["reached"])
        percent = (item["reached"] / item["goal"] * 100) if item["goal"] else 0
        weeks.append({**item, "missing": missing, "percent": round(percent, 1)})

    total_commitment = money(sum(item["commitment"] for item in rows))
    total_reached = money(sum(item["reached"] for item in rows))
    total_missing = money(total_commitment - total_reached)
    total_percent = (total_reached / total_commitment * 100) if total_commitment else 0

    return {
        "workbook": LEGACY_EXCEL_PATH.name,
        "lastModified": datetime.fromtimestamp(LEGACY_EXCEL_PATH.stat().st_mtime).strftime(
            "%d/%m/%Y %H:%M"
        ),
        "summary": {
            "commitment": total_commitment,
            "reached": total_reached,
            "missing": total_missing,
            "percent": round(total_percent, 1),
            "weekGoal": weeks[0]["goal"] if weeks else 0,
            "weekRevenue": weeks[0]["reached"] if weeks else 0,
            "weekMissing": weeks[0]["missing"] if weeks else 0,
            "weekPercent": weeks[0]["percent"] if weeks else 0,
            "positiveCount": sum(1 for item in rows if item["missing"] <= 0),
            "pendingCount": sum(1 for item in rows if item["missing"] > 0),
        },
        "rows": sorted(rows, key=lambda item: item["missing"], reverse=True),
        "history": [],
        "weeks": weeks,
    }


def read_sql_panel(panel):
    if os.environ.get("RENDER") or os.environ.get("RENDER_SERVICE_ID") or os.environ.get("PORT"):
        return None
    try:
        return sensum_sql.read_panel(panel)
    except Exception as exc:
        print(f"[SQL] Falha ao ler painel {panel}; usando fallback local: {exc}")
        return None


def read_dashboard_data():
    sql_data = read_sql_panel("sales")
    if sql_data:
        return apply_weekly_goals(sql_data)

    remote_data = read_remote_payload("data.json", weekly_goals=True)
    if remote_data:
        return remote_data

    if DATA_PATH.exists():
        return apply_weekly_goals(json.loads(DATA_PATH.read_text(encoding="utf-8")))

    if find_compromisso_workbook():
        return apply_weekly_goals(read_compromisso_maio_data())

    if LEGACY_EXCEL_PATH.exists():
        return apply_weekly_goals(read_legacy_excel_dashboard_data())

    raise FileNotFoundError(f"Arquivo nao encontrado: {DATA_PATH}")


def read_positivacao_milho_data():
    sql_data = read_sql_panel("milho")
    if sql_data:
        return sql_data

    remote_data = read_remote_payload("positivacao_milho.json")
    if remote_data:
        return remote_data

    if POSITIVACAO_MILHO_DATA_PATH.exists():
        return json.loads(POSITIVACAO_MILHO_DATA_PATH.read_text(encoding="utf-8"))

    excel_path = find_compromisso_workbook()
    if excel_path:
        workbook = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
        try:
            if "Planilha1" in workbook.sheetnames:
                return read_positivacao_milho_planilha1_data(excel_path, workbook)
        finally:
            workbook.close()

    raise FileNotFoundError(f"Arquivo nao encontrado: {POSITIVACAO_MILHO_DATA_PATH}")


def read_general_data():
    sql_data = read_sql_panel("general")
    if sql_data:
        return sql_data

    remote_data = read_remote_payload("geral.json")
    if remote_data:
        return remote_data

    if GERAL_DATA_PATH.exists():
        return json.loads(GERAL_DATA_PATH.read_text(encoding="utf-8"))

    excel_path = find_compromisso_workbook()
    if excel_path:
        workbook = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
        try:
            if "Planilha1" in workbook.sheetnames:
                return read_general_planilha1_data(excel_path, workbook)
        finally:
            workbook.close()

    raise FileNotFoundError(f"Arquivo nao encontrado: {GERAL_DATA_PATH}")


def read_keys_data():
    sql_data = read_sql_panel("keys")
    if sql_data:
        return sql_data

    remote_data = read_remote_payload("keys.json")
    if remote_data:
        return remote_data

    if KEYS_DATA_PATH.exists():
        return json.loads(KEYS_DATA_PATH.read_text(encoding="utf-8"))

    excel_path = find_compromisso_workbook()
    if excel_path:
        workbook = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
        try:
            if "Planilha1" in workbook.sheetnames:
                return read_keys_planilha1_data(excel_path, workbook)
        finally:
            workbook.close()

    raise FileNotFoundError(f"Arquivo nao encontrado: {KEYS_DATA_PATH}")


def read_published_payload(path, *, weekly_goals=False):
    if not path.exists():
        return None
    payload = json.loads(path.read_text(encoding="utf-8"))
    if weekly_goals:
        return apply_weekly_goals(payload)
    return payload


def read_remote_payload(filename, *, weekly_goals=False):
    cache_key = (filename, weekly_goals)
    now = datetime.now().timestamp()
    cached = REMOTE_CACHE.get(cache_key)
    if cached and now - cached["timestamp"] < REMOTE_CACHE_TTL_SECONDS:
        return deepcopy(cached["payload"])

    api_url = f"{REMOTE_API_BASE}/{filename}?ref=main&ts={int(now)}"
    request = Request(
        api_url,
        headers={
            "Accept": "application/vnd.github+json",
            "User-Agent": "acompanhamento-diario",
        },
    )
    try:
        with urlopen(request, timeout=8) as response:
            api_payload = json.loads(response.read().decode("utf-8"))
            encoded = "".join(str(api_payload.get("content", "")).split())
            payload = json.loads(base64.b64decode(encoded).decode("utf-8"))
    except (OSError, URLError, TimeoutError, ValueError, json.JSONDecodeError):
        raw_url = f"{REMOTE_RAW_BASE}/{filename}?ts={int(now)}"
        raw_request = Request(raw_url, headers={"User-Agent": "acompanhamento-diario"})
        try:
            with urlopen(raw_request, timeout=8) as response:
                payload = json.loads(response.read().decode("utf-8"))
        except (OSError, URLError, TimeoutError, json.JSONDecodeError):
            return None
    if weekly_goals:
        payload = apply_weekly_goals(payload)
    REMOTE_CACHE[cache_key] = {"timestamp": now, "payload": deepcopy(payload)}
    return payload


class DashboardHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        parsed = urlparse(self.path)

        if parsed.path == "/":
            self.send_text(INDEX_HTML, "text/html; charset=utf-8")
            return

        if parsed.path in ("/logo_iatagam.svg", "/logo_iatagam_word.svg"):
            if not LOGO_PATH.exists():
                self.send_text(
                    "Logo nao encontrada",
                    "text/plain; charset=utf-8",
                    status=404,
                )
                return

            payload = LOGO_PATH.read_bytes()
            self.send_response(200)
            self.send_header("Content-Type", "image/png")
            self.send_header("Cache-Control", "no-store")
            self.send_header("Content-Length", str(len(payload)))
            self.end_headers()
            self.wfile.write(payload)
            return

        if parsed.path == "/api/sql-status":
            status = sensum_sql.status()
            payload = json.dumps(status, ensure_ascii=False).encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Cache-Control", "no-store")
            self.send_header("Content-Length", str(len(payload)))
            self.end_headers()
            self.wfile.write(payload)
            return

        if parsed.path == "/api/data":
            try:
                data = read_dashboard_data()
                payload = json.dumps(data, ensure_ascii=False).encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(payload)))
                self.end_headers()
                self.wfile.write(payload)
            except Exception as exc:
                self.send_text(str(exc), "text/plain; charset=utf-8", status=500)
            return
        if parsed.path == "/api/positivacao-milho":
            try:
                data = read_positivacao_milho_data()
                payload = json.dumps(data, ensure_ascii=False).encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(payload)))
                self.end_headers()
                self.wfile.write(payload)
            except Exception as exc:
                self.send_text(str(exc), "text/plain; charset=utf-8", status=500)
            return
        if parsed.path == "/api/geral":
            try:
                data = read_general_data()
                payload = json.dumps(data, ensure_ascii=False).encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(payload)))
                self.end_headers()
                self.wfile.write(payload)
            except Exception as exc:
                self.send_text(str(exc), "text/plain; charset=utf-8", status=500)
            return
        if parsed.path == "/api/keys":
            try:
                data = read_keys_data()
                payload = json.dumps(data, ensure_ascii=False).encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(payload)))
                self.end_headers()
                self.wfile.write(payload)
            except Exception as exc:
                self.send_text(str(exc), "text/plain; charset=utf-8", status=500)
            return
        self.send_text("Nao encontrado", "text/plain; charset=utf-8", status=404)

    def do_POST(self):
        parsed = urlparse(self.path)
        if parsed.path == "/api/weekly-goals":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                body = self.rfile.read(length).decode("utf-8") if length else "{}"
                payload = json.loads(body)
                goals = save_weekly_goals(payload.get("goals", {}))
                response = json.dumps({"goals": goals}, ensure_ascii=False).encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(response)))
                self.end_headers()
                self.wfile.write(response)
            except Exception as exc:
                self.send_text(str(exc), "text/plain; charset=utf-8", status=500)
            return
        self.send_text("Nao encontrado", "text/plain; charset=utf-8", status=404)

    def log_message(self, format, *args):
        timestamp = datetime.now().strftime("%H:%M:%S")
        print(f"[{timestamp}] {self.address_string()} {format % args}")

    def send_text(self, text, content_type, status=200):
        data = text.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)


def main():
    os.chdir(BASE_DIR)
    server = ThreadingHTTPServer((HOST, PORT), DashboardHandler)
    print("Dashboard compartilhado.")
    for address in get_local_addresses():
        print(f"Acesse: {address}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nEncerrando dashboard.")
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
